"""
    Import all libraries bibliotecas what i am using at the script
"""
from warnings import catch_warnings, simplefilter
from openpyxl import load_workbook
from unidecode import unidecode
from pandas import read_excel
from xlwings import Book
from csv import reader

class Excel:
    def __init__(self, pathSpreadsheet: str = None):
        self._pathSpreadsheet = pathSpreadsheet

    '''
        Function of receive values and write at spreadsheet.
        nameSheet -> It must be passs the name of sheet gor paste values.
        columnOne and too much parameters ->  At will be receive the name of columns.
        valuesOne and too much parameters -> At will be receive the values column per column in order.
        with warnings.catch_warnings() -> crate one context at the which is possible manipulation alss errors case show up.
        warnings -> This function will ignor alls erros  de fucntions what he can show up per cause of futures cause.
        Open the spreadsheet 
        cont the last line 
        iterate about the the columns and lines for past values 
        Save spreasheet 
    '''
    def receive_values(self, nameSheet: str = None,
                       columnOne: str = None,
                       columnTwo: str = None,
                       columnThree: str = None,
                       columnfour: str = None,
                       columnFive: str = None,
                       columnSix: str = None,
                       columnSeven: str = None,
                       columnEight: str = None,
                       valuesOne=None,
                       valuesTwo=None,
                       valuesThree=None,
                       valuesFour=None,
                       valuesFive=None,
                       valuesSix=None,
                       valuesSeven=None,
                       valuesEight=None):
    
        with catch_warnings():
            simplefilter(action='ignore', category=FutureWarning)

            self._df = read_excel(self._pathSpreadsheet, sheet_name=nameSheet, index_col=None)

            self._lastLine = self._df.shape[0]

            for i in range(len(valuesThree)):

                if valuesOne:
                    self._df.at[self._lastLine + i, columnOne] = valuesOne

                if valuesTwo:
                    self._df.at[self._lastLine + i, columnTwo] = valuesTwo[i]

                if valuesThree:
                    self._df.at[self._lastLine + i, columnThree] = valuesThree[i]

                if valuesFour:
                    self._df.at[self._lastLine + i, columnfour] = valuesFour

                if valuesFive:
                    self._df.at[self._lastLine + i, columnFive] = valuesFive

                if valuesSix:
                    self._df.at[self._lastLine + i, columnSix] = valuesSix[i]

                if valuesSeven:
                    self._df.at[self._lastLine + i, columnSeven] = valuesSeven[i]

                if valuesEight:
                    self._df.at[self._lastLine + i, columnEight] = valuesEight

            self._df.to_excel(self._pathSpreadsheet, sheet_name=nameSheet, index=False)

    '''
            Function of delete at spreadsheet.
            nameSheet -> It must be pass the name of sheet gor delete values.
            columnOne and too much parameters ->  At will be receive the name of columns for delete values.
            valuesOne and too much parameters -> At will be receive the values column per column for to know if delete or no.
            with warnings.catch_warnings() -> crate one context at the which is possible manipulation alss errors case show up.
            warnings -> This function will ignor alls erros  de fucntions what he can show up per cause of futures cause.
            Open the spreadsheet 
            cont the last line 
            iterate about the the columns and lines for delete values 
            Save spreasheet 
    '''
    def delete_data(self, nameSheet: str = None):

        with catch_warnings():
            simplefilter(action='ignore', category=FutureWarning)

            self._df = read_excel(self._pathSpreadsheet, sheet_name= nameSheet, index_col=None)


            self._lastLine = self._df.shape[0]

            for i in range(self._lastLine):
                for column in self._df.columns:
                    self._df.loc[i, column] = None

            self._df.to_excel(self._pathSpreadsheet, sheet_name=nameSheet, index=False)

    def delete_data_v2(self, nameSheet: str = None):

        self._workbook = load_workbook(self._pathSpreadsheet)

        self._sheet = self._workbook[nameSheet]

        # Apague os dados a partir da linha 2
        for row in self._sheet.iter_rows(min_row=2, max_row=self._sheet.max_row, max_col=self._sheet.max_column):
            for cell in row:
                cell.value = None

        # Salve as alterações no arquivo
        self._workbook.save(self._pathSpreadsheet)

    '''
        Function of run the macro of spreadsheet.
    '''
    def macro(self, module: str = None, sub: str = None):
        self._workbook = Book(self._pathSpreadsheet)
        macr = self._workbook.macro(f'{module}.{sub}')
        macr()

    def write_many_values(self, nameSheet: str = None,
                      columnOne=None,
                      columnTwo=None,
                      columnThree=None,
                      columnFour=None,
                      columnFive=None,
                      columnSix=None,
                      valuesOne=None,
                      valuesTwo=None,
                      valuesThree=None,
                      valuesFour=None,
                      valuesFive=None,
                      valuesSix=None):
        """
        Writes multiple values to a specified Excel sheet in specified columns.

        Args:
            nameSheet (str): The name of the sheet where values will be written.
            columnOne (str): The column letter for the first set of values.
            columnTwo (str): The column letter for the second set of values.
            columnThree (str): The column letter for the third set of values.
            columnFour (str): The column letter for the fourth set of values.
            columnFive (str): The column letter for the fifth set of values.
            columnSix (str): The column letter for the sixth set of values.
            valuesOne (list): The values to write in columnOne.
            valuesTwo (list): The values to write in columnTwo.
            valuesThree (list): The values to write in columnThree.
            valuesFour (list): The values to write in columnFour.
            valuesFive (list): The values to write in columnFive.
            valuesSix (list): The values to write in columnSix.

        Raises:
            Exception: Raises an exception if the provided sheet name is invalid or
                        if there is an error during writing to the spreadsheet.
        """
        
        # Load the existing workbook
        workbook = load_workbook(self._pathSpreadsheet)
        
        # Access the specified worksheet by name
        ws = workbook[nameSheet]

        # Determine the last row in the specified worksheet
        lastLine = ws.max_row + 1

        # Iterate over the range of values in the third column (valuesThree)
        for index in range(0, len(valuesThree)):
            
            # Write values to the specified columns if provided
            if valuesOne:
                ws[columnOne + str(index + lastLine)] = valuesOne
            
            if valuesTwo:
                ws[columnTwo + str(index + lastLine)] = valuesTwo[index]

            if valuesThree:
                ws[columnThree + str(index + lastLine)] = valuesThree[index]

            if valuesFour:
                ws[columnFour + str(index + lastLine)] = valuesFour

            if valuesFive:
                ws[columnFive + str(index + lastLine)] = valuesFive

            if valuesSix:
                ws[columnSix + str(index + lastLine)] = valuesSix[index]

        # Save the changes made to the workbook
        workbook.save(self._pathSpreadsheet)
        
        # Close the workbook to free up resources
        workbook.close()

    def write_many_values_teste(self, nameSheet: str = None,
                             columnOne=None,
                             columnTwo=None,
                             columnThree=None,
                             columnFour=None,
                             columnFive=None,
                             columnSix=None,
                             valuesOne=None,
                             valuesTwo=None,
                             valuesThree=None,
                             valuesFour=None,
                             valuesFive=None,
                             valuesSix=None):
        """
        Writes multiple values to a specified Excel sheet in designated columns,
        ensuring that empty column designations are not written to.

        Args:
            nameSheet (str): The name of the sheet where values will be written.
            columnOne (str): The column letter for the first set of values.
            columnTwo (str): The column letter for the second set of values.
            columnThree (str): The column letter for the third set of values.
            columnFour (str): The column letter for the fourth set of values.
            columnFive (str): The column letter for the fifth set of values.
            columnSix (str): The column letter for the sixth set of values.
            valuesOne (list): The values to write in columnOne.
            valuesTwo (list): The values to write in columnTwo.
            valuesThree (list): The values to write in columnThree.
            valuesFour (list): The values to write in columnFour.
            valuesFive (list): The values to write in columnFive.
            valuesSix (list): The values to write in columnSix.

        Raises:
            Exception: Raises an exception if the provided sheet name is invalid or
                        if there is an error during writing to the spreadsheet.
        """
        
        # Load the existing workbook
        workbook = load_workbook(self._pathSpreadsheet)
        
        # Access the specified worksheet by name
        ws = workbook[nameSheet]

        # Determine the last row in the specified worksheet
        lastLine = ws.max_row + 1

        # Iterate over the range of values in the third column (valuesThree)
        for index in range(0, len(valuesThree)):
            
            # Write values to the specified columns only if they are provided and not empty
            if valuesOne and columnOne != '':
                ws[columnOne + str(index + lastLine)] = valuesOne
            
            if valuesTwo and columnTwo != '':
                ws[columnTwo + str(index + lastLine)] = valuesTwo[index]

            if valuesThree and columnThree != '':
                ws[columnThree + str(index + lastLine)] = valuesThree[index]

            if valuesFour and columnFour != '':
                ws[columnFour + str(index + lastLine)] = valuesFour

            if valuesFive and columnFive != '':
                ws[columnFive + str(index + lastLine)] = valuesFive

            if valuesSix and columnSix != '':
                ws[columnSix + str(index + lastLine)] = valuesSix[index]

        # Save the changes made to the workbook
        workbook.save(self._pathSpreadsheet)

    def insert_values(self, nameSheetOne: str = None, nameSheetTwo: str = None):
        """
        Inserts values from one Excel sheet into another, applying specific transformations
        to the values in the first column.

        Args:
            nameSheetOne (str): The name of the first sheet from which values will be read.
            nameSheetTwo (str): The name of the second sheet from which values will be referenced.

        Raises:
            Exception: Raises an exception if the provided sheet names are invalid or if
                        there is an error during writing to the spreadsheet.
        """
        
        # Load the existing workbook
        self._workBook = load_workbook(self._pathSpreadsheet)
        
        # Access the specified sheets by name
        self._sheet = self._workBook[nameSheetOne]

        # Get the number of active rows in the first sheet
        self._num_linhas_ativas = self._sheet.max_row

            
        # Transform values in the first column (A) to remove accents
        try:
            for index in range(1, self._num_linhas_ativas):
                self._sheet[f'A{index}'].value = unidecode.unidecode(self._sheet[f'A{index}'].value)

        except Exception as e:
            print(f"Error while processing column A: {e}")
        
        # Insert value from cell A1 of the second sheet into column C of the first sheet
        try:
            for row in range(2, self._num_linhas_ativas + 1):
                self._sheet[f'C{row}'] = self._sheetFor['A1'].value
        except Exception as e:
            print(f"Error while inserting value into column C: {e}")
        
        # Insert value from cell B1 of the second sheet into column D of the first sheet
        try:
            for row in range(2, self._num_linhas_ativas + 1):
                self._sheet[f'D{row}'] = self._sheetFor['B1'].value
        except Exception as e:
            print(f"Error while inserting value into column D: {e}")

        # Save the changes made to the workbook
        self._workBook.save(self._pathSpreadsheet)

    
    def read_values(self, nameSheet: str = None,
                columnOne=None,
                columnTwo=None,
                columnThree=None,
                columnFour=None,
                columnFive=None,
                columnSix=None):
        """
        Reads values from specified columns of a given Excel sheet and stores them in a dictionary.

        Args:
            nameSheet (str): The name of the sheet from which values will be read.
            columnOne (str): The column letter for the first set of values.
            columnTwo (str): The column letter for the second set of values.
            columnThree (str): The column letter for the third set of values.
            columnFour (str): The column letter for the fourth set of values.
            columnFive (str): The column letter for the fifth set of values.
            columnSix (str): The column letter for the sixth set of values.

        Returns:
            dict: A dictionary containing lists of values from each specified column, 
                where the key is the column name and the value is a list of cell values.

        Raises:
            Exception: Raises an exception if the provided sheet name is invalid or 
                        if there is an error during reading from the spreadsheet.
        """
        
        # Load the existing workbook
        workbook = load_workbook(self._pathSpreadsheet)
        
        # Access the specified worksheet by name
        ws = workbook[nameSheet]

        # Get the last row of the specified worksheet
        self._last_line = ws.max_row

        # Initialize a dictionary to hold the results from each specified column
        self._result = {
            'columnOne': [],
            'columnTwo': [],
            'columnThree': [],
            'columnFour': [],
            'columnFive': [],
            'columnSix': []
        }
        
        # Iterate over the rows of the specified columns starting from the second row
        for index in range(2, self._last_line + 1):
            
            # Read values from the specified columns and append to the result if they are not None
            if columnOne:
                cell_value = ws[columnOne + str(index)].value
                if cell_value is not None:  # Check if there is data in the cell
                    self._result['columnOne'].append(cell_value)

            if columnTwo:
                cell_value = ws[columnTwo + str(index)].value
                if cell_value is not None:
                    self._result['columnTwo'].append(cell_value)

            if columnThree:
                cell_value = ws[columnThree + str(index)].value
                if cell_value is not None:
                    self._result['columnThree'].append(cell_value)

            if columnFour:
                cell_value = ws[columnFour + str(index)].value
                if cell_value is not None:
                    self._result['columnFour'].append(cell_value)

            if columnFive:
                cell_value = ws[columnFive + str(index)].value
                if cell_value is not None:
                    self._result['columnFive'].append(cell_value)

            if columnSix:
                cell_value = ws[columnSix + str(index)].value
                if cell_value is not None:
                    self._result['columnSix'].append(cell_value)

        # Return the values from the columns that contain data
        return self._result
    
    def csv(self, path, mode, newLine, encode, deli):
        """
        Reads a CSV file and stores its content in a list of tuples.

        Args:
            path (str): The file path to the CSV file.
            mode (str): The mode in which the file is opened (e.g., 'r' for reading).
            newLine (str): Controls how universal newlines work (use '' for universal newlines).
            encode (str): The encoding to use for the file (e.g., 'utf-8').
            deli (str): The delimiter used in the CSV file (e.g., ',' for comma).

        Returns:
            list: A list of tuples, where each tuple represents a row in the CSV file.

        Raises:
            FileNotFoundError: If the specified file path does not exist.
            Exception: Raises an exception for other errors during file reading.
        """
        
        self._lst = []  # Initialize an empty list to store the rows

        # Open the specified CSV file with the provided parameters
        with open(path, mode=mode, newline=newLine, encoding=encode) as file:
            # Create a CSV reader object with the specified delimiter
            self._csv_reader = reader(file, delimiter=deli)
            
            # Iterate through the rows in the CSV file
            for row in self._csv_reader:
                self._lst.append(tuple(row))  # Append each row as a tuple to the list
        
        return self._lst  # Return the list of tuples

    def load_for_xtract(self, name_sheet: str = None, delimiter: str = None, occurences: int = None, length: int = None):
       

        def extract_text(cell_value: str = None, delimiter: str = None, occurences: int = None, length: int = None):
           
            if not isinstance(cell_value, str):
                return ""  # Retorna vazio para células não-textuais
            
            try:
                # Localizar o delimitador
                pos = -1
                for _ in range(occurences):
                    pos = cell_value.find(delimiter, pos + 1)
                    if pos == -1:
                        return ""  # Não encontrou delimitador suficiente
                    
                    
                return cell_value[pos + 1 : pos + 1 + length]  # Retorna o texto limitado
            except Exception as e:
                return f"Erro: {e}"
        
        self.__workbook = load_workbook(self._pathSpreadsheet)
        self.__sheet = self.__workbook[name_sheet]       

        for row in self.__sheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Coluna B
            self.__cell = row[0]
            #print(extract_text(self.__cell.value))
            self.__result = extract_text(self.__cell.value, delimiter, occurences, length)
            self.__sheet[f"C{self.__cell.row}"] = self.__result 

        self.__workbook.save(self._pathSpreadsheet)
        self.__workbook.close()

    def remove_formula(self, name_sheet: str = None):
        # Abrir a planilha
        self.__wb = load_workbook(fr'{self._pathSpreadsheet}', data_only=True)

        # Selecionar a primeira planilha
        self.__sheet = self.__wb[name_sheet]
        
        # Iterar sobre todas as células e substituir as fórmulas pelos valores
        for row in self.__sheet.iter_rows():
            for cell in row:
                # Verificar se a célula tem fórmula
                    cell.value = cell.value  # Substitui a fórmula pelo valor calculado

        # Salvar a planilha com as fórmulas removidas
        self.__wb.save(fr'{self._pathSpreadsheet}')