from openpyxl import load_workbook
from unidecode import unidecode
from .Txt import convertFile
from xlwings import Book, App
from pandas import read_excel
from re import sub, findall

# Defining a class that inherits from 'convertFile'
class Excel_folder_(convertFile):

    # Constructor of the class
    def __init__(self):
        # Calls the constructor of the base class 'convertFile'
        super().__init__()
    
    # Method to open a spreadsheet and process data
    def openSpreadsheet(self, path, nameSheet: str = None, pathTxt: str = None, columnSelect: str = None, destiny_folder: str = None):
        """
        Reads an Excel file, processes its data based on a selected column, 
        and renames text files using the processed data.

        Args:
            path (str): Path to the Excel file.
            nameSheet (str, optional): Name of the sheet to process. Defaults to None.
            pathTxt (str, optional): Path to the text file. Defaults to None.
            columnSelect (str, optional): Name of the column to process. Defaults to None.
            destiny_folder (str, optional): Path to the destination folder. Defaults to None.
        """
        # Stores the path of the spreadsheet and other parameters
        self._path = path
        self._nameSheet = nameSheet
        self._pathTxt = pathTxt

        # Reads the Excel spreadsheet using pandas
        self._workbook = read_excel(self._path)

        # Iterates over the rows of the spreadsheet
        for index, row in self._workbook.iterrows():
            # Processes the specified column
            if columnSelect == 'LocalDeEstocagem':
                # Replaces 'backup-gravacoes' with 'gravacoes'
                self._selection = row[columnSelect].replace('backup-gravacoes', 'gravacoes')
                
                # Fixes redundant network paths
                self._selection = sub(r'\\172\.18\.1\.58\\172\.18\.1\.58\\', r'\\172.18.1.58\\', self._selection)
                
                # Extracts parts of the path using regex
                self._selection = zip(findall(r'172\.18\.1\.58\\(.*)', self._selection))
                
                # Assigns the first match to _selection
                for select in self._selection:
                   self._selection = select[0]
                
            else:
                # If the column is not 'LocalDeEstocagem', simply selects the value   
                self._selection = row[columnSelect]
            
            # Calls the method '_renameDataTxt' to process the text file
            self._renameDataTxt(self._pathTxt, self._selection, columnSelect, destiny_folder)

    # Method to initialize and execute a macro in an .xlsm spreadsheet
    def _initMacro(self):
        """
        Executes a macro in an Excel workbook and closes it.
        """
        # Opens the .xlsm file using xlwings
        self._workbook = Book(r"C:\Users\kaique-ramos\Desktop\Rôbos Fora da Rede\SAI\Main (Recuperado).xlsm")
        # Accesses the macro "Data_files.deleteFomrulas"
        self._m = self._workbook.macro("Data_files.deleteFomrulas")
        # Executes the macro
        self._m()
        # Closes the workbook
        self._workbook.close()
    
    # Method to insert values into a spreadsheet
    def _insertValues(self, pathSpreadsheet, nameSheetOne, nameSheetTwo):
        """
        Inserts data into a spreadsheet from another sheet.

        Args:
            pathSpreadsheet (str): Path to the Excel file.
            nameSheetOne (str): Name of the sheet to insert data into.
            nameSheetTwo (str): Name of the sheet to fetch data from.
        """
        # Loads the specified Excel file
        self._workBook = load_workbook(pathSpreadsheet)
        # Selects the specified sheets
        self._sheet = self._workBook[nameSheetOne]
        self._sheetFor = self._workBook[nameSheetTwo]

        # Gets the number of active rows in 'nameSheetOne'
        self._num_linhas_ativas = self._sheet.max_row

        # Converts the values in column 'A' to a non-accented format
        for index in range(1, self._num_linhas_ativas + 1):
            self._sheet[f'A{index}'].value = unidecode(self._sheet[f'A{index}'].value)

        # Inserts values from 'nameSheetTwo' into 'nameSheetOne'
        for row in range(2, self._num_linhas_ativas + 1):
            self._sheet[f'C{row}'] = self._sheetFor['A1'].value
            self._sheet[f'D{row}'] = self._sheetFor['B1'].value

        # Saves the changes to the Excel file
        self._workBook.save(pathSpreadsheet)

    def formula(self, path: str = None, name_sheet: str = None, path_for_move: str = None):    
        """
        Inserts a formula into a specified Excel sheet using xlwings.

        Args:
            path (str, optional): Path to the Excel file. Defaults to None.
            name_sheet (str, optional): Name of the sheet to insert formulas into. Defaults to None.
        """
        app = App(visible=False)
        # Opens the workbook using xlwings
        try:
            self.__wb = Book(fr'{path}')

            # Selects the specified sheet
            self.__ws = self.__wb.sheets[name_sheet]

            # Gets the number of rows used
            self.__number_lines = self.__ws.range('A2').end('down').row

            # Sets a reference path in column 'E'
            self.__ws.range('E1').value = fr"{path_for_move}"

            # Inserts formulas into column 'D'
            for index in range(2, self.__number_lines + 1):
                self.__ws.range(f'D{index}').formula = (
                    f'=CONCATENATE("move", " ", """", E1, "\\", C{index}, """", " ", """", E1, "\\", """", A{index}, """")'
                )
                
            # Saves the workbook
            self.__wb.save()
            
        finally:
            # Close the workbook and quit the Excel application
            if 'self.__wb' in locals():
                self.__wb.close()
            app.quit() 
        
